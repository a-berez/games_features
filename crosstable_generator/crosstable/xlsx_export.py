from __future__ import annotations

from typing import Sequence

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from .constants import CELL_SEP, DIAGONAL
from .matrix import format_position
from .models import SheetData, TeamResult


def format_cell_xlsx(values: tuple[int, int, int, int]) -> str:
    return CELL_SEP.join(str(value) for value in values)


def set_text_cell(ws: Worksheet, row: int, column: int, value: str | float) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    if isinstance(value, str) and value.startswith("="):
        cell.number_format = "@"


def write_sheet(ws: Worksheet, teams: Sequence[TeamResult], matrix: Sequence[Sequence]) -> None:
    ws.title = ws.title[:31]
    set_text_cell(ws, 1, 1, "место")
    set_text_cell(ws, 1, 2, "команда")
    for col, team_b in enumerate(teams, start=3):
        set_text_cell(ws, 1, col, team_b.name)

    for row, team_a in enumerate(teams, start=2):
        set_text_cell(ws, row, 1, format_position(team_a.position))
        set_text_cell(ws, row, 2, team_a.name)

        for col, cell in enumerate(matrix[row - 2], start=3):
            if cell is None:
                value: str | float = DIAGONAL
            else:
                value = format_cell_xlsx(cell)
            set_text_cell(ws, row, col, value)


def build_workbook(teams: Sequence[TeamResult], sheets: Sequence[SheetData]) -> Workbook:
    workbook = Workbook()
    first = True
    for sheet_data in sheets:
        if first:
            ws = workbook.active
            ws.title = sheet_data.key
            first = False
        else:
            ws = workbook.create_sheet(title=sheet_data.key)
        write_sheet(ws, teams, sheet_data.matrix)
    return workbook
